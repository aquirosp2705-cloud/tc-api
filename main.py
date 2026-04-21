from datetime import datetime
from pathlib import Path
from typing import Literal, List

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook

ARCHIVO_EXCEL = r"C:\MISDOCS\Flujo_MZO26_Final_TC.xlsx"

HOJAS_TARJETAS = {
    "Pichincha": "Pichincha",
    "Produbanco": "Produbanco",
}

ENCABEZADOS = [
    "ID",
    "Fecha",
    "Concepto_Resumen",
    "Concepto_Detallado",
    "Tipo_Consumo",
    "Numero_Cuotas_Pagadas",
    "Numero_Cuotas_Por_Pagar",
    "Valor_Cuota",
    "Ajuste",
    "Pago",
    "Saldo",
    "Saldo_Diferido",
]

CONCEPTOS = [
    "Supermercado",
    "Restaurant",
    "Transporte",
    "Medicinas",
    "Doctor",
    "Salud",
    "Compras exterior",
    "Tecnología",
    "Hogar",
    "Servicios",
    "Ropa",
    "Otros",
]

TIPOS_CONSUMO = ["Corriente", "Diferido"]


class MovimientoIn(BaseModel):
    tarjeta: Literal["Pichincha", "Produbanco"]
    fecha: str = Field(default_factory=lambda: datetime.today().strftime("%Y-%m-%d"))
    concepto_resumen: str
    concepto_detallado: str = ""
    tipo_consumo: Literal["Corriente", "Diferido"] = "Corriente"
    numero_cuotas_pagadas: float = 0
    numero_cuotas_por_pagar: float = 0
    valor_cuota: float = 0
    ajuste: float = 0
    pago: float = 0


class MovimientoOut(BaseModel):
    id: str
    fecha: str
    concepto_resumen: str
    concepto_detallado: str
    tipo_consumo: str
    numero_cuotas_pagadas: float
    numero_cuotas_por_pagar: float
    valor_cuota: float
    ajuste: float
    pago: float
    saldo: float
    saldo_diferido: float


app = FastAPI(title="API Consumos TC", version="1.0.0")


def preparar_archivo():
    path = Path(ARCHIVO_EXCEL)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = HOJAS_TARJETAS["Pichincha"]
        ws.append(ENCABEZADOS)

        ws2 = wb.create_sheet(HOJAS_TARJETAS["Produbanco"])
        ws2.append(ENCABEZADOS)

        wb.save(path)
        wb.close()
        return

    wb = load_workbook(path)
    for hoja in HOJAS_TARJETAS.values():
        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws.append(ENCABEZADOS)
    wb.save(path)
    wb.close()


def obtener_ws(nombre_tarjeta: str):
    preparar_archivo()
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb[HOJAS_TARJETAS[nombre_tarjeta]]
    return wb, ws


def convertir_a_float(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        txt = str(valor).replace("$", "").replace(",", "").strip()
        return float(txt) if txt else 0.0
    except Exception:
        return 0.0


def fila_tiene_datos_usuario(ws, fila: int) -> bool:
    columnas_clave = [2, 3, 4, 5, 10]
    for col in columnas_clave:
        val = ws.cell(row=fila, column=col).value
        if val not in (None, ""):
            return True
    return False


def ultima_fila_real(ws) -> int:
    for fila in range(ws.max_row, 1, -1):
        if fila_tiene_datos_usuario(ws, fila):
            return fila
    return 1


def siguiente_id(ws) -> str:
    contador = max(ultima_fila_real(ws) - 1, 0) + 1
    return f"TC-{contador:05d}"


def saldo_anterior(ws) -> float:
    fila_ant = ultima_fila_real(ws)
    if fila_ant >= 2:
        return convertir_a_float(ws.cell(row=fila_ant, column=11).value)
    return 0.0


def recalcular_desde(ws, fila_inicio: int):
    ultima = ultima_fila_real(ws)
    if fila_inicio < 2:
        fila_inicio = 2

    saldo_corriente = 0.0
    if fila_inicio > 2:
        saldo_corriente = convertir_a_float(ws.cell(row=fila_inicio - 1, column=11).value)

    for fila in range(fila_inicio, ultima + 1):
        pago = convertir_a_float(ws.cell(row=fila, column=10).value)
        tipo = str(ws.cell(row=fila, column=5).value or "").strip()
        valor_cuota = convertir_a_float(ws.cell(row=fila, column=8).value)
        cuotas_por_pagar = convertir_a_float(ws.cell(row=fila, column=7).value)
        ajuste = convertir_a_float(ws.cell(row=fila, column=9).value)

        saldo_corriente += pago + ajuste
        ws.cell(row=fila, column=11, value=saldo_corriente)
        ws.cell(row=fila, column=12, value=valor_cuota * cuotas_por_pagar if tipo == "Diferido" else 0)


def row_to_dict(ws, fila: int) -> dict:
    fecha = ws.cell(fila, 2).value
    if isinstance(fecha, datetime):
        fecha = fecha.strftime("%Y-%m-%d")

    return {
        "id": str(ws.cell(fila, 1).value or ""),
        "fecha": str(fecha or ""),
        "concepto_resumen": str(ws.cell(fila, 3).value or ""),
        "concepto_detallado": str(ws.cell(fila, 4).value or ""),
        "tipo_consumo": str(ws.cell(fila, 5).value or ""),
        "numero_cuotas_pagadas": convertir_a_float(ws.cell(fila, 6).value),
        "numero_cuotas_por_pagar": convertir_a_float(ws.cell(fila, 7).value),
        "valor_cuota": convertir_a_float(ws.cell(fila, 8).value),
        "ajuste": convertir_a_float(ws.cell(fila, 9).value),
        "pago": convertir_a_float(ws.cell(fila, 10).value),
        "saldo": convertir_a_float(ws.cell(fila, 11).value),
        "saldo_diferido": convertir_a_float(ws.cell(fila, 12).value),
    }


@app.get("/")
def root():
    return {
        "ok": True,
        "mensaje": "API Consumos TC funcionando",
        "tarjetas": list(HOJAS_TARJETAS.keys()),
    }


@app.get("/tarjetas")
def tarjetas():
    return {"tarjetas": list(HOJAS_TARJETAS.keys())}


@app.get("/movimientos/{tarjeta}", response_model=List[MovimientoOut])
def listar_movimientos(tarjeta: Literal["Pichincha", "Produbanco"]):
    wb, ws = obtener_ws(tarjeta)
    ultima = ultima_fila_real(ws)
    datos = [row_to_dict(ws, fila) for fila in range(2, ultima + 1)]
    wb.close()
    return datos


@app.get("/saldo/{tarjeta}")
def saldo_tarjeta(tarjeta: Literal["Pichincha", "Produbanco"]):
    wb, ws = obtener_ws(tarjeta)
    saldo = saldo_anterior(ws)
    ultima = ultima_fila_real(ws)
    saldo_diferido = convertir_a_float(ws.cell(row=ultima, column=12).value) if ultima >= 2 else 0.0
    wb.close()
    return {
        "tarjeta": tarjeta,
        "saldo": saldo,
        "saldo_diferido_ultima_linea": saldo_diferido,
        "ultima_fila": ultima,
    }


@app.post("/movimientos")
def crear_movimiento(mov: MovimientoIn):
    if mov.concepto_resumen not in CONCEPTOS:
        raise HTTPException(status_code=400, detail="Concepto_Resumen no válido.")

    wb, ws = obtener_ws(mov.tarjeta)

    cuotas_pagadas = mov.numero_cuotas_pagadas
    cuotas_por_pagar = mov.numero_cuotas_por_pagar
    valor_cuota = mov.valor_cuota
    ajuste = mov.ajuste

    if mov.tipo_consumo == "Corriente":
        cuotas_pagadas = 0
        cuotas_por_pagar = 0
        valor_cuota = 0
        ajuste = 0

    fila = ultima_fila_real(ws) + 1
    saldo_prev = saldo_anterior(ws)

    ws.cell(row=fila, column=1, value=siguiente_id(ws))
    ws.cell(row=fila, column=2, value=mov.fecha)
    ws.cell(row=fila, column=3, value=mov.concepto_resumen)
    ws.cell(row=fila, column=4, value=mov.concepto_detallado)
    ws.cell(row=fila, column=5, value=mov.tipo_consumo)
    ws.cell(row=fila, column=6, value=cuotas_pagadas)
    ws.cell(row=fila, column=7, value=cuotas_por_pagar)
    ws.cell(row=fila, column=8, value=valor_cuota)
    ws.cell(row=fila, column=9, value=ajuste)
    ws.cell(row=fila, column=10, value=mov.pago)
    ws.cell(row=fila, column=11, value=saldo_prev + mov.pago + ajuste)
    ws.cell(row=fila, column=12, value=valor_cuota * cuotas_por_pagar if mov.tipo_consumo == "Diferido" else 0)

    wb.save(ARCHIVO_EXCEL)
    data = row_to_dict(ws, fila)
    wb.close()

    return {"ok": True, "fila": fila, "movimiento": data}


@app.delete("/movimientos/{tarjeta}/{id_registro}")
def eliminar_movimiento(tarjeta: Literal["Pichincha", "Produbanco"], id_registro: str):
    wb, ws = obtener_ws(tarjeta)
    ultima = ultima_fila_real(ws)
    fila_encontrada = None

    for fila in range(2, ultima + 1):
        if str(ws.cell(row=fila, column=1).value or "").strip() == id_registro.strip():
            fila_encontrada = fila
            break

    if fila_encontrada is None:
        wb.close()
        raise HTTPException(status_code=404, detail="No se encontró el registro.")

    ws.delete_rows(fila_encontrada, 1)

    nueva_ultima = ultima_fila_real(ws)
    for fila in range(2, nueva_ultima + 1):
        ws.cell(row=fila, column=1, value=f"TC-{fila-1:05d}")

    recalcular_desde(ws, fila_encontrada)

    wb.save(ARCHIVO_EXCEL)
    wb.close()

    return {"ok": True, "eliminado": id_registro, "tarjeta": tarjeta}
